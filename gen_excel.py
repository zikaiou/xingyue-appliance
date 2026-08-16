# -*- coding: utf-8 -*-
"""Generate XINGYUE product master Excel file."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Master"

# Header style
pink = PatternFill(start_color="FF2F7D", end_color="FF2F7D", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Model", "Category", "Product Name", "Key Features", "Spec Sheet", "Detail Page"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = pink
    c.font = header_font
    c.alignment = header_align
    c.border = border

rows = [
    ["ST-718", "Travel Iron", "Compact Travel Iron", "Compact Travel Design; Fast Heating; Portable Ironing Solution", "https://www.xingyue-appliance.com/specs/st-718-spec.pdf", "https://www.xingyue-appliance.com/products/st-718"],
    ["ST-717", "Travel Iron", "Portable Travel Iron", "Portable Design; Convenient Travel Use; Efficient Ironing Performance", "https://www.xingyue-appliance.com/specs/st-717-spec.pdf", "https://www.xingyue-appliance.com/products/st-717"],
    ["ST-588", "Travel Iron", "Travel Iron", "Compact Structure; Easy Carry Design; Daily Travel Solution", "https://www.xingyue-appliance.com/specs/st-588-spec.pdf", "https://www.xingyue-appliance.com/products/st-588"],
    ["ST-812", "Travel Iron", "Travel Iron", "Lightweight Travel Design; Quick Wrinkle Removal; Portable Garment Care", "https://www.xingyue-appliance.com/specs/st-812-spec.pdf", "https://www.xingyue-appliance.com/products/st-812"],
    ["ST-8807", "Travel Iron", "Professional Travel Iron", "Advanced Steam System; High-Efficiency Heating; Professional Results", "https://www.xingyue-appliance.com/specs/st-8807-spec.pdf", "https://www.xingyue-appliance.com/products/st-8807"],
    ["ST-8808", "Travel Iron", "Travel Iron", "Compact Power Design; Steady Steam Flow; Versatile Ironing", "https://www.xingyue-appliance.com/specs/st-8808-spec.pdf", "https://www.xingyue-appliance.com/products/st-8808"],
    ["ST-817", "Travel Iron", "Travel Iron", "Digital Control Interface; Enhanced Steam Output; Premium Fabric Care", "https://www.xingyue-appliance.com/specs/st-817-spec.pdf", "https://www.xingyue-appliance.com/products/st-817"],
    ["ST-200A", "Travel Iron", "Travel Iron", "Ergonomic Handle; Multi-Angle Steaming; Travel Safety Build", "https://www.xingyue-appliance.com/specs/st-200a-spec.pdf", "https://www.xingyue-appliance.com/products/st-200a"],
    ["ST-815", "Garment Steamer", "Handheld Garment Steamer", "1800W Powerful Steam; 260ml Tank; Vertical & Horizontal", "https://www.xingyue-appliance.com/specs/st-815-spec.pdf", "https://www.xingyue-appliance.com/products/st-815"],
    ["6617", "Garment Steamer", "Garment Steamer", "Ceramic Steam Panel; 3 Steam Levels; Auto Power Off", "https://www.xingyue-appliance.com/specs/6617-spec.pdf", "https://www.xingyue-appliance.com/products/6617"],
    ["900A", "Garment Steamer", "Garment Steamer", "Stainless Steel Panel; 20±5g/min Steam; Compact Portable", "https://www.xingyue-appliance.com/specs/900a-spec.pdf", "https://www.xingyue-appliance.com/products/900a"],
    ["6618", "Garment Steamer", "Garment Steamer", "Stainless Steel Panel; 3 Steam Modes; Travel Friendly", "https://www.xingyue-appliance.com/specs/6618-spec.pdf", "https://www.xingyue-appliance.com/products/6618"],
    ["ST-S1001", "Steam Iron", "Steam Iron", "2800W Powerful Steam; 330ml Tank; Auto Shut-Off", "https://www.xingyue-appliance.com/specs/st-s1001-spec.pdf", "https://www.xingyue-appliance.com/products/st-s1001"],
    ["ST-S1002", "Steam Iron", "Steam Iron", "3000W Powerful Steam; 420ml Tank; Auto Shut-Off", "https://www.xingyue-appliance.com/specs/st-s1002-spec.pdf", "https://www.xingyue-appliance.com/products/st-s1002"],
    ["ST-S1003", "Steam Iron", "Steam Iron", "2800W; 330ml Tank; Ceramic/Non-stick Soleplate", "https://www.xingyue-appliance.com/specs/st-s1003-spec.pdf", "https://www.xingyue-appliance.com/products/st-s1003"],
    ["ST-S1004", "Steam Iron", "Steam Iron", "2800W; 330ml Tank; Self-Clean System", "https://www.xingyue-appliance.com/specs/st-s1004-spec.pdf", "https://www.xingyue-appliance.com/products/st-s1004"],
    ["9002", "Vacuum Garment Steamer", "Vacuum Garment Steamer", "1500W Vacuum Steam; 400ml Tank; LED Display", "https://www.xingyue-appliance.com/specs/9002-spec.pdf", "https://www.xingyue-appliance.com/products/9002"],
    ["9003", "Vacuum Garment Steamer", "3-in-1 Vacuum Garment Steamer", "3-in-1 Dry/Steam/Vacuum; 1500W; Dual-Level", "https://www.xingyue-appliance.com/specs/9003-spec.pdf", "https://www.xingyue-appliance.com/products/9003"],
    ["9004", "Vacuum Garment Steamer", "3-in-1 Vacuum Garment Steamer", "3-in-1; Rotatable Head; 1300W", "https://www.xingyue-appliance.com/specs/9004-spec.pdf", "https://www.xingyue-appliance.com/products/9004"],
    ["9005", "Vacuum Garment Steamer", "Vacuum Garment Steamer", "1500W; 400ml Tank; LED Display & Ceramic Plate", "https://www.xingyue-appliance.com/specs/9005-spec.pdf", "https://www.xingyue-appliance.com/products/9005"],
]

row_font = Font(size=10)
link_font = Font(size=10, color="FF2F7D")
for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = link_font if c >= 5 else row_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, 4)))

# Column widths
widths = [14, 22, 28, 50, 55, 55]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(rows)+1}"

out = r"C:\Users\asus\AccioWork\2026-07-17-14-23-26\XINGYUE_Product_Master.xlsx"
wb.save(out)
print("Saved:", out)
